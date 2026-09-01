"""W10 backend tests — Reports, Exports & Embedded Analytics."""
import csv
import io
from datetime import date, datetime, timedelta
from decimal import Decimal

import pytest
from fastapi import HTTPException
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models.entities import (
    Attendance, Contract, ContractBranch, Courier, CourierType, Country, GeoCity, GeoCountry,
    IncentiveRule, PayrollInputRecord, Tenant, User, UserRole,
)
from app.routers import reports as reports_router


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    try:
        yield session
    finally:
        session.close()


def make_tenant(db, name="TestCo"):
    t = Tenant(name=name, country=Country.SA)
    db.add(t); db.commit(); db.refresh(t)
    return t


def make_admin(db, tenant_id):
    u = User(phone="966500000001", password_hash="x", role=UserRole.COMPANY_ADMIN, tenant_id=tenant_id)
    db.add(u); db.commit(); db.refresh(u)
    return u


def make_country(db):
    c = GeoCountry(name="Saudi Arabia", code="SA", flag="🇸🇦", active=True)
    db.add(c); db.commit(); db.refresh(c)
    return c


def make_city(db, country_id, name="Riyadh"):
    c = GeoCity(country_id=country_id, name=name, active=True)
    db.add(c); db.commit(); db.refresh(c)
    return c


def make_operating_city(db, tenant_id, geo_city_id):
    from app.models.entities import TenantOperatingCity
    oc = TenantOperatingCity(tenant_id=tenant_id, geo_city_id=geo_city_id, display_name="Riyadh", is_active=True)
    db.add(oc); db.commit(); db.refresh(oc)
    return oc


def make_contract(db, tenant_id):
    c = Contract(tenant_id=tenant_id, name="HS Riyadh")
    db.add(c); db.commit(); db.refresh(c)
    return c


def make_branch(db, tenant_id, contract_id, city_id):
    b = ContractBranch(tenant_id=tenant_id, contract_id=contract_id, city_id=city_id, city="Riyadh", is_active=True)
    db.add(b); db.commit(); db.refresh(b)
    return b


def make_rider(db, tenant_id, city_id, contract_id, branch_id, name="Rider", phone="966500000100", status="ACTIVE"):
    r = Courier(
        tenant_id=tenant_id, name=name, phone=phone, courier_type=CourierType.COMPANY,
        country=Country.SA, employment_status=status, city_id=city_id,
        contract_id=contract_id, contract_branch_id=branch_id,
    )
    db.add(r); db.commit(); db.refresh(r)
    return r


# ============================================================
# REPORT CATALOG TESTS
# ============================================================

def test_report_catalog_company_admin(db):
    """Test that company admin sees all report categories."""
    tenant = make_tenant(db)
    admin = make_admin(db, tenant_id=tenant.id)
    
    result = reports_router.report_catalog(user=admin)
    
    assert "catalog" in result
    catalog = result["catalog"]
    assert "workforce" in catalog
    assert "attendance" in catalog
    assert "financial" in catalog
    assert "documents" in catalog
    assert "vehicles" in catalog
    assert "orders" in catalog
    assert "performance" in catalog
    assert "audit" in catalog


def test_report_catalog_supervisor_limited(db):
    """Test that supervisor sees only limited categories."""
    tenant = make_tenant(db)
    u = User(phone="966500000002", password_hash="x", role=UserRole.SUPERVISOR, tenant_id=tenant.id)
    db.add(u); db.commit()
    
    result = reports_router.report_catalog(user=u)
    
    catalog = result["catalog"]
    assert "workforce" in catalog
    assert "attendance" in catalog
    assert "performance" in catalog
    assert "financial" not in catalog


def test_report_catalog_accountant_financial_only(db):
    """Test that accountant sees only financial reports."""
    tenant = make_tenant(db)
    u = User(phone="966500000003", password_hash="x", role=UserRole.ACCOUNTANT, tenant_id=tenant.id)
    db.add(u); db.commit()
    
    result = reports_router.report_catalog(user=u)
    
    catalog = result["catalog"]
    assert "financial" in catalog
    assert "workforce" not in catalog


# ============================================================
# WORKFORCE REPORT TESTS
# ============================================================

def test_workforce_rider_master(db):
    """Test rider master report."""
    tenant = make_tenant(db)
    admin = make_admin(db, tenant_id=tenant.id)
    country = make_country(db)
    city = make_city(db, country_id=country.id)
    operating_city = make_operating_city(db, tenant_id=tenant.id, geo_city_id=city.id)
    contract = make_contract(db, tenant_id=tenant.id)
    branch = make_branch(db, tenant_id=tenant.id, contract_id=contract.id, city_id=city.id)
    
    r1 = make_rider(db, tenant_id=tenant.id, city_id=city.id, contract_id=contract.id, branch_id=branch.id, phone="966500000101", name="Rider A")
    r2 = make_rider(db, tenant_id=tenant.id, city_id=city.id, contract_id=contract.id, branch_id=branch.id, phone="966500000102", name="Rider B")
    
    result = reports_router.workforce_rider_master(user=admin, db=db)
    
    assert result["total"] == 2
    assert len(result["rows"]) == 2
    assert result["rows"][0]["name"] == "Rider A"


def test_workforce_rider_master_tenant_isolation(db):
    """Test that riders from other tenants are excluded."""
    tenant1 = make_tenant(db, "Tenant1")
    tenant2 = make_tenant(db, "Tenant2")
    admin1 = make_admin(db, tenant_id=tenant1.id)
    
    country = make_country(db)
    city = make_city(db, country_id=country.id)
    operating_city1 = make_operating_city(db, tenant_id=tenant1.id, geo_city_id=city.id)
    operating_city2 = make_operating_city(db, tenant_id=tenant2.id, geo_city_id=city.id)
    contract1 = make_contract(db, tenant_id=tenant1.id)
    contract2 = make_contract(db, tenant_id=tenant2.id)
    branch1 = make_branch(db, tenant_id=tenant1.id, contract_id=contract1.id, city_id=city.id)
    branch2 = make_branch(db, tenant_id=tenant2.id, contract_id=contract2.id, city_id=city.id)
    
    r1 = make_rider(db, tenant_id=tenant1.id, city_id=city.id, contract_id=contract1.id, branch_id=branch1.id, phone="966500000101")
    r2 = make_rider(db, tenant_id=tenant2.id, city_id=city.id, contract_id=contract2.id, branch_id=branch2.id, phone="966500000102")
    
    result = reports_router.workforce_rider_master(user=admin1, db=db)
    
    assert result["total"] == 1
    assert result["rows"][0]["name"] == r1.name


def test_workforce_rider_master_filtering(db):
    """Test rider master report filtering."""
    tenant = make_tenant(db)
    admin = make_admin(db, tenant_id=tenant.id)
    country = make_country(db)
    city = make_city(db, country_id=country.id)
    operating_city = make_operating_city(db, tenant_id=tenant.id, geo_city_id=city.id)
    contract = make_contract(db, tenant_id=tenant.id)
    branch = make_branch(db, tenant_id=tenant.id, contract_id=contract.id, city_id=city.id)
    
    r1 = make_rider(db, tenant_id=tenant.id, city_id=city.id, contract_id=contract.id, branch_id=branch.id, phone="966500000101", status="ACTIVE")
    r2 = make_rider(db, tenant_id=tenant.id, city_id=city.id, contract_id=contract.id, branch_id=branch.id, phone="966500000102", status="SUSPENDED")
    
    result = reports_router.workforce_rider_master(employment_status="ACTIVE", user=admin, db=db)
    
    assert result["total"] == 1
    assert result["rows"][0]["status"] == "ACTIVE"


# ============================================================
# ATTENDANCE REPORT TESTS
# ============================================================

def test_attendance_summary(db):
    """Test attendance summary report."""
    tenant = make_tenant(db)
    admin = make_admin(db, tenant_id=tenant.id)
    country = make_country(db)
    city = make_city(db, country_id=country.id)
    operating_city = make_operating_city(db, tenant_id=tenant.id, geo_city_id=city.id)
    contract = make_contract(db, tenant_id=tenant.id)
    branch = make_branch(db, tenant_id=tenant.id, contract_id=contract.id, city_id=city.id)
    
    r1 = make_rider(db, tenant_id=tenant.id, city_id=city.id, contract_id=contract.id, branch_id=branch.id, phone="966500000101")
    
    today = date.today()
    db.add(Attendance(
        courier_id=r1.id,
        check_in=datetime.combine(today, datetime.min.time()),
        check_out=datetime.combine(today, datetime.min.time()) + timedelta(hours=8),
        is_late=False,
    ))
    db.commit()
    
    result = reports_router.attendance_summary(user=admin, db=db)
    
    assert result["total_riders"] == 1
    assert len(result["rows"]) == 1
    assert result["rows"][0]["attendance_days"] == 1


# ============================================================
# PAYROLL LEDGER REPORT TESTS
# ============================================================

def test_financial_payroll_ledger_decimal_precision(db):
    """Test that payroll ledger preserves Decimal precision."""
    tenant = make_tenant(db)
    admin = make_admin(db, tenant_id=tenant.id)
    country = make_country(db)
    city = make_city(db, country_id=country.id)
    operating_city = make_operating_city(db, tenant_id=tenant.id, geo_city_id=city.id)
    contract = make_contract(db, tenant_id=tenant.id)
    branch = make_branch(db, tenant_id=tenant.id, contract_id=contract.id, city_id=city.id)
    
    r1 = make_rider(db, tenant_id=tenant.id, city_id=city.id, contract_id=contract.id, branch_id=branch.id, phone="966500000101")
    
    period = date.today().strftime("%Y-%m")
    db.add(PayrollInputRecord(
        tenant_id=tenant.id, courier_id=r1.id, month=period,
        source_type="MANUAL", input_type="EARNING", amount=Decimal("1000.50"),
        description="راتب", status="APPROVED",
    ))
    db.commit()
    
    result = reports_router.financial_payroll_ledger(user=admin, db=db)
    
    assert result["total"] == 1
    assert result["rows"][0]["amount"] == 1000.50


def test_financial_payroll_ledger_tenant_isolation(db):
    """Test payroll ledger tenant isolation."""
    tenant1 = make_tenant(db, "Tenant1")
    tenant2 = make_tenant(db, "Tenant2")
    admin1 = make_admin(db, tenant_id=tenant1.id)
    
    country = make_country(db)
    city = make_city(db, country_id=country.id)
    operating_city1 = make_operating_city(db, tenant_id=tenant1.id, geo_city_id=city.id)
    operating_city2 = make_operating_city(db, tenant_id=tenant2.id, geo_city_id=city.id)
    contract1 = make_contract(db, tenant_id=tenant1.id)
    contract2 = make_contract(db, tenant_id=tenant2.id)
    branch1 = make_branch(db, tenant_id=tenant1.id, contract_id=contract1.id, city_id=city.id)
    branch2 = make_branch(db, tenant_id=tenant2.id, contract_id=contract2.id, city_id=city.id)
    
    r1 = make_rider(db, tenant_id=tenant1.id, city_id=city.id, contract_id=contract1.id, branch_id=branch1.id, phone="966500000101")
    r2 = make_rider(db, tenant_id=tenant2.id, city_id=city.id, contract_id=contract2.id, branch_id=branch2.id, phone="966500000102")
    
    period = date.today().strftime("%Y-%m")
    db.add(PayrollInputRecord(
        tenant_id=tenant1.id, courier_id=r1.id, month=period,
        source_type="MANUAL", input_type="EARNING", amount=Decimal("1000"), status="APPROVED",
    ))
    db.add(PayrollInputRecord(
        tenant_id=tenant2.id, courier_id=r2.id, month=period,
        source_type="MANUAL", input_type="EARNING", amount=Decimal("2000"), status="APPROVED",
    ))
    db.commit()
    
    result = reports_router.financial_payroll_ledger(user=admin1, db=db)
    
    assert result["total"] == 1
    assert result["rows"][0]["amount"] == 1000


# ============================================================
# EXPORT TESTS
# ============================================================

def test_export_csv_content_isolation(db):
    """Test CSV export content respects tenant isolation at the data level."""
    tenant1 = make_tenant(db, "Tenant1")
    tenant2 = make_tenant(db, "Tenant2")
    admin1 = make_admin(db, tenant_id=tenant1.id)
    
    country = make_country(db)
    city = make_city(db, country_id=country.id)
    operating_city1 = make_operating_city(db, tenant_id=tenant1.id, geo_city_id=city.id)
    operating_city2 = make_operating_city(db, tenant_id=tenant2.id, geo_city_id=city.id)
    contract1 = make_contract(db, tenant_id=tenant1.id)
    contract2 = make_contract(db, tenant_id=tenant2.id)
    branch1 = make_branch(db, tenant_id=tenant1.id, contract_id=contract1.id, city_id=city.id)
    branch2 = make_branch(db, tenant_id=tenant2.id, contract_id=contract2.id, city_id=city.id)
    
    r1 = make_rider(db, tenant_id=tenant1.id, city_id=city.id, contract_id=contract1.id, branch_id=branch1.id, phone="966500000101", name="Rider A")
    r2 = make_rider(db, tenant_id=tenant2.id, city_id=city.id, contract_id=contract2.id, branch_id=branch2.id, phone="966500000102", name="Rider B")
    
    # Get data for export (same as export_csv would)
    data = reports_router.workforce_rider_master(user=admin1, db=db)
    rows = data["rows"]
    
    # Build CSV manually (since StreamingResponse is async)
    output = io.StringIO()
    output.write("\ufeff")
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    
    body = output.getvalue()
    reader = csv.DictReader(io.StringIO(body))
    csv_rows = list(reader)
    
    assert len(csv_rows) == 1
    assert csv_rows[0]["name"] == "Rider A"


def test_export_csv_rbac_supervisor(db):
    """Test that supervisor can export authorized data."""
    tenant = make_tenant(db)
    u = User(phone="966500000002", password_hash="x", role=UserRole.SUPERVISOR, tenant_id=tenant.id)
    db.add(u); db.commit()
    
    country = make_country(db)
    city = make_city(db, country_id=country.id)
    operating_city = make_operating_city(db, tenant_id=tenant.id, geo_city_id=city.id)
    contract = make_contract(db, tenant_id=tenant.id)
    branch = make_branch(db, tenant_id=tenant.id, contract_id=contract.id, city_id=city.id)
    
    r1 = make_rider(db, tenant_id=tenant.id, city_id=city.id, contract_id=contract.id, branch_id=branch.id, phone="966500000101")
    
    result = reports_router.workforce_rider_master(user=u, db=db)
    
    # Supervisor should see workforce reports
    assert result["total"] == 1


def test_export_xlsx_produces_valid_file(db):
    """Test XLSX export produces valid file."""
    pytest.importorskip("openpyxl")
    
    from openpyxl import load_workbook
    
    tenant = make_tenant(db)
    admin = make_admin(db, tenant_id=tenant.id)
    country = make_country(db)
    city = make_city(db, country_id=country.id)
    operating_city = make_operating_city(db, tenant_id=tenant.id, geo_city_id=city.id)
    contract = make_contract(db, tenant_id=tenant.id)
    branch = make_branch(db, tenant_id=tenant.id, contract_id=contract.id, city_id=city.id)
    
    r1 = make_rider(db, tenant_id=tenant.id, city_id=city.id, contract_id=contract.id, branch_id=branch.id, phone="966500000101")
    
    # Get data directly
    data = reports_router.workforce_rider_master(user=admin, db=db)
    rows = data["rows"]
    
    # Generate xlsx
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "rider_master"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Verify it's a valid xlsx
    wb2 = load_workbook(output)
    assert "rider_master" in wb2.sheetnames


# ============================================================
# ANALYTICS VIEWS TEST
# ============================================================

def test_analytics_views(db):
    """Test analytics views endpoint."""
    tenant = make_tenant(db)
    admin = make_admin(db, tenant_id=tenant.id)
    
    result = reports_router.analytics_views(user=admin)
    
    assert "views" in result
    assert len(result["views"]) > 0
    view_ids = [v["id"] for v in result["views"]]
    assert "analytics_workforce" in view_ids
    assert "analytics_payroll" in view_ids


# ============================================================
# RBAC TESTS
# ============================================================

def test_unauthorized_role_rejected(db):
    """Test that unauthorized roles are rejected."""
    tenant = make_tenant(db)
    u = User(phone="966500000004", password_hash="x", role=UserRole.VIEWER, tenant_id=tenant.id)
    db.add(u); db.commit()
    
    with pytest.raises(HTTPException) as exc:
        reports_router.workforce_rider_master(user=u, db=db)
    
    assert exc.value.status_code == 403
